"""Hedged-debt lookups for the IRS Valuation feed (col AW, Hedged Debt MTM).

A trade's ``hedge`` direction decides where AW comes from:

* ``Long``  -> the MTM of the debt the swap hedges, taken as
  ``Clean + USD Outstanding``. The chain is::

      IRS quantum_deal_number
          -> Debt Deal Number          (data/debt/Deal_Numbers.csv, static map)
          -> Clean + USD Outstanding   (data/debt/Deal_Summary_<val_date>.xlsx)

* ``Short`` -> the swap's own clean value (``v.clean``); no debt files needed.

Original signs are preserved for Long; Short reverses the swap clean's sign.

Files (under ``data/debt/`` by default):

* ``Deal_Numbers.csv`` -- header ``Debt Deal Number,IRS Deal Number``.
* ``Deal_Summary_<val_date>.xlsx`` -- Sheet1 with a free-form title in row 1,
  column headers in row 2, data from row 3. We read ``Debt Deal Number``,
  ``Clean`` and ``USD Outstanding`` by header name (the file carries many
  other columns).
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

import openpyxl

DEAL_NUMBERS_FILE = "Deal_Numbers.csv"


def _norm_deal(x: object) -> str:
    """Normalize a deal number to a bare string key.

    Deal numbers arrive as ints (xlsx), strings (CSV/trade rows), or pandas
    floats (``19085763.0``). Strip whitespace and a spurious trailing ``.0`` so
    every source keys identically."""
    if x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def debt_summary_filename(val_date: date) -> str:
    """Spec filename: ``Deal_Summary_<YYYY-MM-DD>.xlsx``."""
    return f"Deal_Summary_{val_date.isoformat()}.xlsx"


def load_deal_number_map(path: str | Path) -> dict[str, str]:
    """``IRS Deal Number -> Debt Deal Number`` from ``Deal_Numbers.csv``.

    Raises:
        FileNotFoundError: file missing.
        ValueError:        required header columns absent.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Deal-number map not found: {p}")
    with p.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return {}
    header = [c.strip() for c in rows[0]]
    try:
        debt_i = header.index("Debt Deal Number")
        irs_i = header.index("IRS Deal Number")
    except ValueError:
        raise ValueError(
            f"{p}: header must contain 'Debt Deal Number' and 'IRS Deal "
            f"Number'; got {header}"
        )
    out: dict[str, str] = {}
    for r in rows[1:]:
        if not any((c or "").strip() for c in r):
            continue
        irs = _norm_deal(r[irs_i]) if irs_i < len(r) else ""
        debt = _norm_deal(r[debt_i]) if debt_i < len(r) else ""
        if irs:
            out[irs] = debt
    return out


def load_debt_mtm(path: str | Path) -> dict[str, float]:
    """``Debt Deal Number -> Clean + USD Outstanding`` from a ``Deal_Summary``.

    The Long-hedge MTM (col AW) is the debt's ``Clean`` plus its ``USD
    Outstanding`` notional, so this loader reads both columns and returns
    their sum per deal.

    Layout: row 1 free-form title, row 2 column headers, row 3+ data.

    Raises:
        FileNotFoundError: file missing.
        ValueError:        required header columns absent.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Debt summary not found: {p}")
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if len(rows) < 2:
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[1]]
    try:
        debt_i = header.index("Debt Deal Number")
        clean_i = header.index("Clean")
        out_i = header.index("USD Outstanding")
    except ValueError:
        raise ValueError(
            f"{p}: header row 2 must contain 'Debt Deal Number', 'Clean' and "
            f"'USD Outstanding'; got {header}"
        )
    out: dict[str, float] = {}
    for r in rows[2:]:
        if r is None or max(debt_i, clean_i, out_i) >= len(r):
            continue
        deal = _norm_deal(r[debt_i])
        if not deal or r[clean_i] is None or r[out_i] is None:
            continue
        try:
            out[deal] = float(r[clean_i]) + float(r[out_i])
        except (TypeError, ValueError):
            continue
    return out


def resolve_hedged_debt_mtm(
    trade_id: str,
    hedge: str,
    quantum_deal_number: str,
    swap_clean: float,
    deal_map: dict[str, str],
    debt_mtm: dict[str, float],
) -> float:
    """Compute the Hedged Debt MTM (col AW) for one trade.

    ``Short`` -> ``-swap_clean`` (the swap's clean value with its sign
    **reversed**). ``Long`` -> the hedged debt's ``Clean + USD Outstanding``,
    resolved ``quantum_deal_number -> Debt Deal Number -> Clean + USD
    Outstanding`` (sign preserved).

    Raises ``ValueError`` (a hard, per-trade error) when ``hedge`` is blank /
    unrecognized, or a ``Long`` trade cannot be resolved to a debt Clean. The
    Portfolio runner catches it, records the trade in ``manifest.errors[]``, and
    the run ends ``status="partial"``.
    """
    h = (hedge or "").strip().lower()
    if h == "short":
        return -swap_clean
    if h == "long":
        irs_deal = _norm_deal(quantum_deal_number)
        if not irs_deal:
            raise ValueError(
                f"{trade_id}: hedge=Long requires a quantum_deal_number to "
                f"look up the hedged debt's MTM."
            )
        debt_deal = deal_map.get(irs_deal, "")
        if not debt_deal:
            raise ValueError(
                f"{trade_id}: IRS deal number {irs_deal!r} is not mapped to a "
                f"debt deal number in {DEAL_NUMBERS_FILE}."
            )
        if debt_deal not in debt_mtm:
            raise ValueError(
                f"{trade_id}: debt deal number {debt_deal!r} (mapped from IRS "
                f"deal {irs_deal!r}) not found in the Debt Summary."
            )
        return debt_mtm[debt_deal]
    raise ValueError(
        f"{trade_id}: 'hedge' must be 'Long' or 'Short' (got {hedge!r}); it is "
        f"required on every trade row for the IRS Valuation feed."
    )
