"""Excel-based loaders.

Curve file layout (one file per valuation date), normalized to an in-memory
``ZeroCurve`` so nothing downstream changes:

  Raw production format:
    - Filename: ``market_environment_YYYY-MM-DD.csv`` (ISO date, dashes)
    - A few non-data header rows on top (``Name``/``Date``/``Property`` in col A).
    - Many irrelevant curve pillars (other currencies, EQ/FX/VOL tickers) are
      interleaved with the ones we need. They are filtered out automatically by
      ``TICKER_RE`` (col-A content filter) -- only
      ``IR.USD-{SOFR|FEDFUNDS}-ON.ZERORATE-{TENOR}.MID`` rows are kept.
    - Col A: ticker, Col B: zero rate (decimal, e.g. 0.0364).
    - One file holds **both** SOFR and FEDFUNDS curves (interleaved).

Fixings file layout:
  - Excel or CSV with columns ``date`` and ``rate`` (optionally a leading
    ``ticker`` column). Production filename is
    ``fixing_cail_USD-FEDFUNDS-ON.csv``; content is identical to the legacy
    ``fedfunds.csv`` so no special handling is required.
"""

from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from ..curve import ZeroCurve
from ..fixings import FixingHistory
from ..rate_quoting import DEFAULT, RateQuoting
from .base import CurveLoader, FixingLoader

CURVE_RAW_RE = re.compile(
    r"^market_environment[_-](\d{4}-\d{2}-\d{2})\.csv$", re.IGNORECASE
)
TICKER_RE = re.compile(r"^IR\.USD-(SOFR|FEDFUNDS)-ON\.ZERORATE-([0-9A-Z]+)\.MID$")

CANONICAL_NAME = {
    "SOFR": "SOFR",
    "FEDFUNDS": "FEDFUNDS",
    "FF": "FEDFUNDS",
    "FED_FUNDS": "FEDFUNDS",
}


class ExcelCurveLoader(CurveLoader):
    def __init__(
        self,
        base_dir: str | Path,
        rate_quoting: RateQuoting | None = None,
        sheet_name: str = "Sheet1",
    ) -> None:
        self.base_dir = Path(base_dir)
        self.rate_quoting = rate_quoting or DEFAULT
        self.sheet_name = sheet_name
        self._cache: dict[tuple[date, str], dict[str, float]] = {}

    def _file_path(self, val_date: date) -> Path:
        raw_name = f"market_environment_{val_date.strftime('%Y-%m-%d')}.csv"
        p = self.base_dir / raw_name
        if p.exists():
            return p
        # Loose, case-insensitive match (handles market_environment-<date>.csv too).
        for f in self.base_dir.iterdir():
            if not f.is_file():
                continue
            m = CURVE_RAW_RE.match(f.name)
            if m and m.group(1) == val_date.strftime("%Y-%m-%d"):
                return f
        raise FileNotFoundError(
            f"Curve file not found for {val_date} in {self.base_dir} "
            f"(expected {raw_name})"
        )

    def _iter_ticker_rows(self, path: Path) -> Iterator[tuple[str, object]]:
        """Yield ``(col_a, col_b)`` for every data row, regardless of file type.

        Non-data header rows and irrelevant pillars are not filtered here --
        callers apply ``TICKER_RE`` so the same content filter is shared by the
        legacy xlsx and raw csv paths.
        """
        if path.suffix.lower() == ".csv":
            with path.open("r", newline="", encoding="utf-8-sig") as fh:
                for row in csv.reader(fh):
                    if not row or not row[0].strip() or len(row) < 2:
                        continue
                    yield row[0], row[1]
        else:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                ws = self._select_ws(wb)
                for row in ws.iter_rows(values_only=True):
                    if not row or row[0] is None or len(row) < 2:
                        continue
                    yield row[0], row[1]
            finally:
                wb.close()

    def _select_ws(self, wb: "openpyxl.workbook.workbook.Workbook"):
        if self.sheet_name in wb.sheetnames:
            return wb[self.sheet_name]
        if "in" in wb.sheetnames:  # raw export sheet name
            return wb["in"]
        return wb[wb.sheetnames[0]]

    def _parse_all_pillars(self, val_date: date) -> dict[str, dict[str, float]]:
        path = self._file_path(val_date)
        pillars: dict[str, dict[str, float]] = {"SOFR": {}, "FEDFUNDS": {}}
        for col_a, col_b in self._iter_ticker_rows(path):
            m = TICKER_RE.match(str(col_a).strip())
            if not m:
                continue
            index, tenor = m.group(1), m.group(2)
            pillars[index][tenor] = float(str(col_b).strip())
        if not pillars["SOFR"] or not pillars["FEDFUNDS"]:
            raise ValueError(f"Curve file {path} missing SOFR or FEDFUNDS pillars")
        return pillars

    def load(self, val_date: date, curve_name: str) -> ZeroCurve:
        key = CANONICAL_NAME.get(curve_name.upper(), curve_name.upper())
        if (val_date, "ALL") not in self._cache:
            parsed = self._parse_all_pillars(val_date)
            self._cache[(val_date, "SOFR")] = parsed["SOFR"]
            self._cache[(val_date, "FEDFUNDS")] = parsed["FEDFUNDS"]
            self._cache[(val_date, "ALL")] = {}  # marker
        pillars = self._cache.get((val_date, key))
        if pillars is None:
            raise ValueError(f"Unknown curve {curve_name!r}; known: SOFR, FEDFUNDS")
        return ZeroCurve(val_date, pillars, self.rate_quoting, name=key)

    def load_from_file(self, path: str | Path, val_date: date, curve_name: str) -> ZeroCurve:
        """Load directly from an explicit file path, bypassing the base_dir/filename convention.

        Useful for ad-hoc inspection of curve files that live outside ``data/curves/``.
        """
        key = CANONICAL_NAME.get(curve_name.upper(), curve_name.upper())
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Curve file not found: {path}")
        pillars: dict[str, float] = {}
        for col_a, col_b in self._iter_ticker_rows(path):
            m = TICKER_RE.match(str(col_a).strip())
            if not m:
                continue
            index, tenor = m.group(1), m.group(2)
            if index == key:
                pillars[tenor] = float(str(col_b).strip())
        if not pillars:
            raise ValueError(f"No {key} pillars found in {path}")
        return ZeroCurve(val_date, pillars, self.rate_quoting, name=key)


class ExcelFixingLoader(FixingLoader):
    """Load historical fixings from a CSV/XLSX file.

    Accepted layouts (auto-detected):
      * 2-column ``date, rate`` -- with or without a header row.
      * 3-column ``ticker, date, rate`` -- with or without a header row.
        When a ticker column is present, rows whose ticker contains
        ``index_name`` (case-insensitive, hyphens/underscores ignored) are kept;
        if no rows match, all rows are kept.

    Date parsing is flexible: ISO (YYYY-MM-DD), US (M/D/YYYY), or anything
    pandas can coerce via ``to_datetime``.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def load(self, index_name: str) -> FixingHistory:
        if not self.path.exists():
            return FixingHistory({}, name=index_name)

        if self.path.suffix.lower() == ".csv":
            raw = pd.read_csv(self.path, header=None, skip_blank_lines=True, dtype=str)
        else:
            raw = pd.read_excel(self.path, header=None, dtype=str)
        raw = raw.dropna(how="all").reset_index(drop=True)
        if raw.empty:
            return FixingHistory({}, name=index_name)

        # Detect header by attempting to parse the last column of the first row as a number
        try:
            float(str(raw.iloc[0, -1]).strip())
            has_header = False
        except (ValueError, TypeError):
            has_header = True
        if has_header:
            raw = raw.iloc[1:].reset_index(drop=True)

        ncols = raw.shape[1]
        if ncols < 2:
            raise ValueError(f"Fixings file {self.path} needs >=2 columns (date,rate); got {ncols}")
        if ncols == 2:
            ticker_col, date_col, rate_col = None, 0, 1
        else:
            ticker_col, date_col, rate_col = 0, 1, 2

        # Filter by ticker substring match, if applicable
        if ticker_col is not None and index_name:
            norm = index_name.upper().replace("_", "").replace("-", "")
            mask = (
                raw[ticker_col].astype(str).str.upper()
                .str.replace("_", "", regex=False)
                .str.replace("-", "", regex=False)
                .str.contains(norm, na=False)
            )
            if mask.any():
                raw = raw[mask].reset_index(drop=True)

        mapping: dict[date, float] = {}
        for _, row in raw.iterrows():
            d_raw, r_raw = row[date_col], row[rate_col]
            if pd.isna(d_raw) or pd.isna(r_raw):
                continue
            d = pd.to_datetime(str(d_raw).strip(), errors="coerce")
            if pd.isna(d):
                continue
            try:
                mapping[d.date()] = float(str(r_raw).strip())
            except (ValueError, TypeError):
                continue
        return FixingHistory(mapping, name=index_name)
