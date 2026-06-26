"""Item 7: Day 1 Valuations (frequency Once, Email).

Excel per ``Day 1 Valuations.xlsx``: a summary block, a fixed-vs-floating leg
summary, then per-leg cashflow detail (as in the debug workbook). Produced only
for swap id(s) named via ``--new-deal-<id>``.

ASSUMPTIONS (confirm — see _intake.md):
* Key Rate = par rate; Total Value = clean + accrued.
* PV01 = present value of a 1bp fixed-leg annuity; split DV01 computed per leg
  (fixed + floating = total). Floating PV01 / Spot Exchange / Cash Accrued left
  blank (0 where the sample shows 0), per the example.
* Per-leg "Clean Price"/"Total Value" = signed leg PV (sign from pay/receive fixed).
* Floating "Forward Rate" = compounded coupon; Fixing Date = period's last fixing.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from .base import RunContext
from .helpers import coupon_rows, period_fixing_dates, xldate, xlnum
from .priced import leg_risk


def _filename(raw_id: str, val_date: date) -> str:
    return f"{raw_id} {val_date:%m.%d.%Y} - Day 1 Valuations.xlsx"


def _signs(pay_fixed: bool) -> tuple[float, float]:
    """(fixed_leg_sign, floating_leg_sign): receive-fixed = +fixed/-float."""
    s = -1.0 if pay_fixed else 1.0
    return s, -s


def _write_sheet(ws, pt, val_date: date, md) -> None:
    td, v = pt.trade, pt.valuation
    fx_sign, fl_sign = _signs(bool(td.pay_fixed))
    total_value = v.clean + v.accrued
    risk = leg_risk(pt, md)

    # 1. Summary block (label, value) in cols A/B.
    summary = [
        ("Key Rate:", xlnum(v.par_rate)),
        ("DV01:", xlnum(v.dv01)),
        ("PV01:", xlnum(risk["pv01"])),
        ("Clean Price:", xlnum(v.clean)),
        ("MTM Accrued Interest:", xlnum(v.accrued)),
        ("Cash Accrued Interest:", 0),
        ("Total Value:", xlnum(total_value)),
        ("Timestamp:", datetime.now()),
        ("Value Date:", val_date),
        ("Valuation Currency", "USD"),
    ]
    row = 1
    for label, value in summary:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # 2. Fixed vs Floating leg summary (A/B and I/J).
    row += 1
    ws.cell(row=row, column=1, value="Fixed Leg Value:")
    ws.cell(row=row, column=9, value="Floating Leg Value:")
    leg_rows = [
        ("DV01:", xlnum(risk["dv01_fixed"]), "DV01:", xlnum(risk["dv01_floating"])),
        ("PV01:", xlnum(risk["pv01"]), "PV01:", None),  # PV01 is the fixed-leg annuity; floating blank per sample
        ("Clean Price:", xlnum(fx_sign * v.pv_fixed), "Clean Price:", xlnum(fl_sign * v.pv_floating)),
        ("MTM Accrued Interest:", 0, "MTM Accrued Interest:", 0),
        ("Cash Accrued Interest:", 0, "Cash Accrued Interest:", 0),
        ("Total Value:", xlnum(fx_sign * v.pv_fixed), "Total Value:", xlnum(fl_sign * v.pv_floating)),
        ("Spot Exchange:", None, "Spot Exchange:", None),
    ]
    for f_lbl, f_val, g_lbl, g_val in leg_rows:
        row += 1
        ws.cell(row=row, column=1, value=f_lbl)
        ws.cell(row=row, column=2, value=f_val)
        ws.cell(row=row, column=9, value=g_lbl)
        ws.cell(row=row, column=10, value=g_val)

    # 3. Per-leg cashflow detail (side by side: fixed cols A-H, floating J-S).
    row += 2
    ws.cell(row=row, column=1, value="Fixed Leg Values")
    ws.cell(row=row, column=10, value="Floating Leg Values")
    row += 1
    fixed_hdr = ["Start Date", "End Date", "Payment Date", "Notional", "Fixed Rate",
                 "Discount Factor", "Cashflow FV", "Cashflow PV"]
    float_hdr = ["Start Date", "End Date", "Fixing Date", "Payment Date", "Notional",
                 "Forward Rate", "Spread", "Discount Factor", "Cashflow FV", "Cashflow PV"]
    for i, h in enumerate(fixed_hdr):
        ws.cell(row=row, column=1 + i, value=h)
    for i, h in enumerate(float_hdr):
        ws.cell(row=row, column=10 + i, value=h)

    fixed = coupon_rows(v.fixed_cf).reset_index(drop=True)
    floating = coupon_rows(v.floating_cf_by_period).reset_index(drop=True)
    fixings = period_fixing_dates(v.floating_cf)
    base = row + 1

    for i, (_, r) in enumerate(fixed.iterrows()):
        vals = [xldate(r["accrual_start"]), xldate(r["accrual_end"]), xldate(r["payment_date"]),
                xlnum(r["notional"]), xlnum(r["coupon_rate"]), xlnum(r["df_to_payment"]),
                xlnum(r["payment_amount"]), xlnum(r["discounted_cashflow"])]
        for j, val in enumerate(vals):
            ws.cell(row=base + i, column=1 + j, value=val)

    for i, (_, r) in enumerate(floating.iterrows()):
        fix = fixings.get((r["accrual_start"], r["accrual_end"]))
        vals = [xldate(r["accrual_start"]), xldate(r["accrual_end"]), xldate(fix),
                xldate(r["payment_date"]), xlnum(r["notional"]), xlnum(r["compounded_coupon"]),
                xlnum(r["spread"]), xlnum(r["df_to_payment"]), xlnum(r["payment_amount"]),
                xlnum(r["discounted_cashflow"])]
        for j, val in enumerate(vals):
            ws.cell(row=base + i, column=10 + j, value=val)


def produce(ctx: RunContext, dest_dir: Path) -> list[Path]:
    pp = ctx.priced()
    written: list[Path] = []
    dest_dir.mkdir(parents=True, exist_ok=True)

    for raw_id in sorted(ctx.new_deal_ids):
        matches = pp.by_raw_id(raw_id)
        if not matches:
            continue
        wb = Workbook()
        ws = wb.active
        ws.title = "Day 1 Valuation"
        # One swap per file; if an id maps to multiple, use the first sheet + extras.
        _write_sheet(ws, matches[0], ctx.val_date, pp.md)
        for extra in matches[1:]:
            _write_sheet(wb.create_sheet(), extra, ctx.val_date, pp.md)
        out = dest_dir / _filename(raw_id, ctx.val_date)
        wb.save(out)
        written.append(out)
    return written
