"""Generate output field mapping Excel files for the writeup."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── style helpers ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")   # mid blue
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")   # light blue
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Calibri", size=10)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="top")

thin = Side(style="thin", color="B8CCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_ws(ws, col_widths):
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_header_row(ws, row, values, fill, font, heights=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = CENTER
        c.border = BORDER
    if heights:
        ws.row_dimensions[row].height = heights


def write_data_row(ws, row, values, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[row].height = None  # auto


def write_section_row(ws, row, label, ncols):
    c = ws.cell(row=row, column=1, value=label)
    c.fill = SUBHDR_FILL
    c.font = SUBHDR_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# VALUATION OUTPUT MAP
# ══════════════════════════════════════════════════════════════════════════════

VAL_COLS = [
    "Output Column Letter",
    "Field Name",
    "Data Type",
    "Source / Derivation",
    "Notes",
]

# (col_letter, field_name, data_type, source, notes)
VAL_ROWS = [
    # ── section ──
    ("SECTION", "TRADE IDENTIFICATION & REFERENCE FIELDS", None, None, None),
    ("A", "Trade Reference Number",    "String",  "Always blank",
     "Not required per KPMG spec."),
    ("B", "Internal Reference Number", "String",  "Always blank",
     "Not required per KPMG spec."),
    ("C", "Quantum Deal Number",       "String",  "Trade definition: td.quantum_deal_number",
     "Unique trade ID in the firm's Quantum system."),
    ("D", "Oracle Entity Code",        "String",  "Trade definition: td.oracle_entity_code",
     "4-digit entity code; also drives CCID construction (cols AU/AV)."),
    ("E", "Notional Currency",         "String",  "Trade definition: td.notional_currency",
     "Always USD for this portfolio."),
    ("F", "As of Date",                "Date",    "CLI argument: val_date",
     "The market/valuation date passed when the pricer is run."),

    # ── section ──
    ("SECTION", "CALCULATED VALUATION OUTPUTS", None, None, None),
    ("G", "Clean Price",               "Numeric", "Pricer: v.clean  =  dirty − accrued",
     "Mark-to-market value excluding accrued interest. Positive = asset."),
    ("H", "Accrued Interest",          "Numeric", "Pricer: v.accrued  (both legs, net of pay/receive direction)",
     "Interest earned in the current period on both legs but not yet paid. "
     "Uses inclusive day-count (val_date itself is counted)."),
    ("I", "Total Value (NPV / Dirty)", "Numeric", "Pricer: v.dirty  =  PV(receive leg) − PV(pay leg)",
     "Sum of all discounted future cash flows. clean + accrued = dirty holds exactly."),
    ("J", "DV01",                      "Numeric", "Pricer: v.dv01  =  PV_base − PV_bumped(+1bp)",
     "Full revaluation: both SOFR and FF curves shifted +1bp in parallel. "
     "Positive DV01 = position loses value when rates rise."),
    ("K", "Valuation Currency",        "String",  "Constant: 'USD'",
     "Hard-coded; book is USD-only."),

    # ── section ──
    ("SECTION", "PERIOD / COMPONENT DETAIL (NOT REQUIRED)", None, None, None),
    ("L", "Child Reference Number",   "String",  "Always blank", "Not required per spec."),
    ("M", "Period Start Date",        "Date",    "Always blank", "Not required per spec."),
    ("N", "Period End Date",          "Date",    "Always blank", "Not required per spec."),
    ("O", "Payment Date",             "Date",    "Always blank", "Not required per spec."),

    # ── section ──
    ("SECTION", "NOTIONAL & TERM", None, None, None),
    ("P", "Maturity Date",            "Date",    "Trade definition: td.maturity_date",
     "The termination / last accrual end date of the swap."),
    ("Q", "Notional 1 Amount",        "Numeric", "Trade definition: td.notional",
     "Face amount in USD. Also summed in the footer row."),
    ("R", "Notional 1 Amount USD",    "Numeric", "Trade definition: td.notional",
     "Same as Q; both carry the same value (USD-only book)."),

    # ── section ──
    ("SECTION", "COMPONENT FAIR VALUE FIELDS (NOT REQUIRED)", None, None, None),
    ("S", "Pay Rec Status",           "String",  "Always blank", "Not required per spec."),
    ("T", "Component Type",           "String",  "Always blank", "Not required per spec."),
    ("U", "Coupon FV",                "Numeric", "Always blank (footer sum = 0)", "Not required per spec."),
    ("V", "Intrinsic Value FV",       "Numeric", "Always blank (footer sum = 0)", "Not required per spec."),
    ("W", "Time Value FV",            "Numeric", "Always blank (footer sum = 0)", "Not required per spec."),

    # ── section ──
    ("SECTION", "COUNTERPARTY & ENTITY REFERENCE FIELDS", None, None, None),
    ("X", "Intercompany Trade",       "String",  "Trade definition: td.intercompany → 'Yes' / 'No'",
     "True/false boolean from trade file rendered as Yes/No."),
    ("Y", "Counterparty Name (Quantum)", "String", "Trade definition: td.counterparty_name_quantum",
     "Counterparty name as recorded in Quantum."),
    ("Z", "Current Counterparty",    "String",  "Trade definition: td.current_counterparty",
     "Legal counterparty name. Exact string 'CME Clearing House' triggers CME-branch logic."),
    ("AA", "Entity Name (Quantum)",  "String",  "Trade definition: td.entity_name_quantum",
     "Amex entity abbreviation with Oracle entity code."),
    ("AB", "Reporting Party",        "String",  "Trade definition: td.reporting_party",
     "Legal name of the Amex reporting entity."),
    ("AC", "InternalFacing-StreetFacing", "String", "Always blank", "Not required per spec."),

    # ── section ──
    ("SECTION", "PRODUCT CLASSIFICATION", None, None, None),
    ("AD", "Product",                "String",  "Constant: 'IR'",
     "Hard-coded; all instruments in this portfolio are interest rate products."),
    ("AE", "Sub-Product2",           "String",
     "CME branch: 'OTC - Centralized (Principal)';  else: 'OTC - Bilateral'",
     "Determined by whether current_counterparty == 'CME Clearing House'."),
    ("AF", "Collateral Level",       "String",  "Constant: 'Fully Collateralized'",
     "Hard-coded per portfolio-level agreement."),
    ("AG", "Counterparty Code",      "String",  "Always blank", "Not required per spec."),
    ("AH", "Counterparty Type",      "String",
     "CME branch: 'Financial Market Utility';  else: 'Bank'",
     "Determined by CME branch logic (same flag as AE)."),
    ("AI", "Counterparty Location",  "String",  "Trade definition: td.counterparty_location",
     "Country / domicile of the counterparty (e.g., 'US', 'UK')."),
    ("AJ", "HCL Type",               "String",  "Constant: 'Interest Rate Swap'",
     "Hard-coded; identifies instrument type for the KPMG feed."),

    # ── section ──
    ("SECTION", "ASSET / LIABILITY SPLIT", None, None, None),
    ("AK", "DA (Derivative Asset)",  "Numeric",
     "NPV if NPV > 0; else blank",
     "Gross asset value. Summed in footer. Blank when NPV ≤ 0."),
    ("AL", "DL (Derivative Liability)", "Numeric",
     "abs(NPV) if NPV < 0; else blank",
     "Gross liability value (positive number). Summed in footer. Blank when NPV ≥ 0."),
    ("AM", "Asset Liability Tag",    "String",
     "NPV > 0 → 'Asset';  NPV < 0 → 'Liability';  NPV = 0 → blank",
     "Classification flag derived from sign of dirty NPV."),

    # ── section ──
    ("SECTION", "CLEARING / CCP FLAGS", None, None, None),
    ("AN", "Qualifying CCP",         "String",
     "CME branch: 'Yes';  else: 'No'", ""),
    ("AO", "Cleared",                "String",
     "CME branch: 'Yes';  else: 'No'", ""),
    ("AP", "Cash-Settled CCP",       "String",
     "CME branch: 'Yes';  else: 'No'", ""),

    # ── section ──
    ("SECTION", "DEAL DATES & NETTING", None, None, None),
    ("AQ", "Deal Date",              "Date",    "Trade definition: td.deal_date",
     "Trade inception / strike date (distinct from effective start_date)."),
    ("AR", "Netting ID",             "String",  "Trade definition: td.netting_id",
     "Key into the Netting Database; groups trades for netting calculations."),
    ("AS", "Cash Flow Netting Allowed", "String",
     "Netting Database lookup by netting_id",
     "Whether cash flows within the netting group may be netted."),
    ("AT", "Position Netting Allowed",  "String",
     "Netting Database lookup by netting_id",
     "Whether positions within the netting group may be netted."),

    # ── section ──
    ("SECTION", "ACCOUNTING IDENTIFIERS (CCID)", None, None, None),
    ("AU", "Balance Sheet CCID",     "String",
     "Composed: oracle_entity_code + Default_RC (Entity_Reference_Report.csv) + "
     "Natural Account (192001 if Asset, 392001 if Liability) + fixed trailing segments",
     "9-segment dash-joined code. Blank if entity lookup fails or NPV = 0. "
     "Format: Entity-RC-NaturalAcct-000000-0000-000000-000000-000000-0000."),
    ("AV", "PL OCI CCID",            "String",
     "Composed: oracle_entity_code + Default_RC + Natural Account 465012 (always, regardless of sign) "
     "+ fixed trailing segments",
     "9-segment dash-joined code. Blank if entity lookup fails. "
     "Natural account 465012 is fixed for P&L / OCI regardless of asset or liability status."),

    # ── section ──
    ("SECTION", "HEDGED DEBT MTM", None, None, None),
    ("AW", "Hedged Debt MTM",        "Numeric",
     "Short hedge (SC): −v.clean (negative of swap's clean value).  "
     "Long hedge (LH): the hedged bond is valued in-process from the trade's inline "
     "debt_* fields (FixedLeg model, principal-at-maturity, discounted on Fed Funds + "
     "debt_discount_spread, signed from the obligor's view), then AW = debt Clean + "
     "USD Outstanding (debt_notional).",
     "Required field. An LH trade whose inline debt cannot be priced is a hard "
     "per-trade error. Legacy external Deal_Numbers.csv / Deal_Summary_<date>.xlsx "
     "feed is no longer used. Summed in footer row."),
]

# ══════════════════════════════════════════════════════════════════════════════
# NETTING OUTPUT MAP
# ══════════════════════════════════════════════════════════════════════════════

NET_COLS = [
    "Output Column Letter",
    "Field Name",
    "Data Type",
    "Source / Derivation",
    "Notes",
]

NET_ROWS = [
    # ── section ──
    ("SECTION", "FILE / RUN IDENTIFICATION", None, None, None),
    ("A", "Field",          "String",  "Constant: 'Position Netting'",
     "Hard-coded label identifying the row type."),
    ("B", "As of Date",     "Date",    "CLI argument: val_date",
     "The valuation date for this run."),
    ("C", "Product",        "String",  "Constant: 'IRS'",
     "Hard-coded product label."),
    ("D", "Entity",         "String",  "Constant: 'American Express Company'",
     "Hard-coded reporting entity name."),

    # ── section ──
    ("SECTION", "NETTING GROUP IDENTIFICATION", None, None, None),
    ("E", "Oracle Entity Code",  "String",
     "Netting Database: netting_db[netting_id].netting_entity",
     "Entity code for the netting group, sourced from the Netting Database."),
    ("F", "Counterparty",        "String",
     "First trade's current_counterparty within the netting group",
     "Counterparty name is taken from the first trade belonging to this netting_id."),
    ("G", "Counterparty Code",   "String",  "Always blank", "Not required per spec."),
    ("H", "Payment Date",        "Date",    "Always blank", "Not required per spec."),
    ("I", "Maturity Date",       "Date",    "Always blank", "Not required per spec."),
    ("J", "Netting ID",          "String",  "Netting group key (from trade definitions)",
     "The netting_id that groups this set of trades together."),

    # ── section ──
    ("SECTION", "GROSS EXPOSURE AMOUNTS", None, None, None),
    ("K", "Gross DA",            "Numeric",
     "Sum of NPV for all trades in the group where NPV > 0",
     "Total gross asset exposure for the netting group. Summed in footer."),
    ("L", "Gross DL",            "Numeric",
     "Sum of abs(NPV) for all trades in the group where NPV < 0",
     "Total gross liability exposure (positive number). Summed in footer."),

    # ── section ──
    ("SECTION", "NETTING CALCULATION", None, None, None),
    ("M", "Netting Amount",      "Numeric",
     "min(Gross DA, Gross DL)  if position_netting_allowed = Yes;  else 0",
     "The amount by which gross exposures are reduced through netting. "
     "Zero if position netting is not contractually permitted for this group. Summed in footer."),
    ("N", "Net DA",              "Numeric",
     "Gross DA − Netting Amount",
     "Residual asset exposure after applying the netting offset. Summed in footer."),
    ("O", "Net DL",              "Numeric",
     "Gross DL − Netting Amount",
     "Residual liability exposure after applying the netting offset. Summed in footer."),

    # ── section ──
    ("SECTION", "COUNTERPARTY CLASSIFICATION & NETTING RULES", None, None, None),
    ("P", "Counterparty Type",       "String",
     "CME branch: 'FMU';  else: 'Bank'",
     "Derived from whether the group's counterparty is 'CME Clearing House'."),
    ("Q", "Cash Flow Netting Allowed",  "String",
     "Netting Database: netting_db[netting_id].cash_flow_netting_allowed",
     "Whether cash flow netting is contractually permitted for this group."),
    ("R", "Position Netting Allowed",   "String",
     "Netting Database: netting_db[netting_id].position_netting_allowed",
     "Whether position netting is contractually permitted (drives the Netting Amount calc)."),
    ("S", "Netting Entity",             "String",
     "Netting Database: netting_db[netting_id].netting_entity",
     "Legal entity name associated with this netting group."),

    # ── section ──
    ("SECTION", "ACCOUNTING IDENTIFIERS (CCID)", None, None, None),
    ("T", "Position Netting Asset CCID",     "String",
     "Composed: netting_entity + RC + 192005 + 000000-0000-000000-000000-000000-0000",
     "9-segment CCID for netting group's asset side. Natural account 192005 is fixed for netting assets."),
    ("U", "Position Netting Liability CCID", "String",
     "Composed: netting_entity + RC + 392004 + 000000-0000-000000-000000-000000-0000",
     "9-segment CCID for netting group's liability side. Natural account 392004 is fixed for netting liabilities."),
]


def build_wb(sheet_title, col_headers, rows, col_widths, title_text):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Title row
    ws.merge_cells("A1:" + get_column_letter(len(col_headers)) + "1")
    tc = ws["A1"]
    tc.value = title_text
    tc.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column header row
    write_header_row(ws, 2, col_headers, SUBHDR_FILL, SUBHDR_FONT, heights=20)

    # Freeze panes at row 3
    ws.freeze_panes = "A3"

    current_row = 3
    alt = False
    for r in rows:
        col_letter, fname, dtype, source, notes = r
        if col_letter == "SECTION":
            write_section_row(ws, current_row, f"  {fname}", len(col_headers))
            current_row += 1
            alt = False
            continue
        write_data_row(ws, current_row,
                       [col_letter, fname, dtype or "", source or "", notes or ""],
                       alt=alt)
        # make the source cell wrap nicely
        ws.cell(current_row, 4).alignment = WRAP
        ws.cell(current_row, 5).alignment = WRAP
        ws.row_dimensions[current_row].height = None
        alt = not alt
        current_row += 1

    style_ws(ws, col_widths)
    return wb


# ── build and save ─────────────────────────────────────────────────────────────

val_wb = build_wb(
    sheet_title="IRS Valuation Outputs",
    col_headers=VAL_COLS,
    rows=VAL_ROWS,
    col_widths=[8, 28, 12, 60, 55],
    title_text="IRS Valuation Output Field Map  —  IRS_Valuation_<val_date>-00001.csv  (49 columns, A–AW)",
)
val_path = os.path.join(OUT_DIR, "IRS_Valuation_Output_Map.xlsx")
val_wb.save(val_path)
print(f"Saved: {val_path}")

net_wb = build_wb(
    sheet_title="IRS Netting Outputs",
    col_headers=NET_COLS,
    rows=NET_ROWS,
    col_widths=[8, 32, 12, 60, 55],
    title_text="IRS Netting Output Field Map  —  IRS_Netting_<val_date>-00001.csv  (21 columns, A–U)",
)
net_path = os.path.join(OUT_DIR, "IRS_Netting_Output_Map.xlsx")
net_wb.save(net_path)
print(f"Saved: {net_path}")
