import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── palette ──────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
YLW_FILL = PatternFill("solid", fgColor="FFF2CC")
GRN_FILL = PatternFill("solid", fgColor="E2EFDA")

HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
SEC_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)
ITAL_FONT = Font(italic=True, name="Calibri", size=9, color="595959")
TTL_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=13)

thin = Side(style="thin", color="B8CCE4")

def tb():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, val):
    cell.value = val
    cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = tb()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sec(cell, val):
    cell.value = val
    cell.font = SEC_FONT; cell.fill = SEC_FILL; cell.border = tb()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def body(cell, val, alt=False, bold=False, fill=None, align="left"):
    cell.value = val
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.fill = fill if fill else (ALT_FILL if alt else WHT_FILL)
    cell.border = tb()
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

def banner(ws, row, span, text):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws.cell(row, 1, text)
    c.font = TTL_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def note(ws, row, span, text):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws.cell(row, 1, text)
    c.font = ITAL_FONT
    c.fill = PatternFill("solid", fgColor="F2F2F2")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 40

def ex_title(ws, row, span, text):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    c = ws.cell(row, 1, text)
    c.font = SEC_FONT; c.fill = SEC_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

# ════════════════════════════════════════════════════════════════════════
# TAB 1  Field Reference
# ════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Field Reference"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

for i, w in enumerate([28, 26, 38, 22, 18, 52], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

banner(ws1, 1, 6, "Swap Pricer — Input Convention Field Reference")
ws1.row_dimensions[2].height = 30
for c, h in enumerate(["Field Name","Plain-English Name","Accepted Values",
                        "Default","Applies To","Business Impact"], 1):
    hdr(ws1.cell(2, c), h)

ROWS = [
    # section, field, name, values, default, applies, impact
    ("SHARED — TRADE-LEVEL ECONOMIC TERMS", None,None,None,None,None),
    (None,"trade_id","Trade Identifier","Any string","—","Both legs",
     "Unique key used across all output files and audit logs"),
    (None,"notional","Notional Amount","Positive number","—","Both legs",
     "Face value of the swap; all cashflows scale linearly with this"),
    (None,"pay_fixed","Direction","true / false","—","Both legs",
     "true = you pay fixed and receive floating (standard payer swap);\nfalse = receive fixed and pay floating (receiver swap)"),
    (None,"fixed_rate","Agreed Fixed Rate","Decimal  e.g. 0.0525 = 5.25%","—","Fixed leg",
     "The contractual coupon rate; set at trade inception and never changes"),
    (None,"start_date","Effective Date","YYYY-MM-DD","—","Both legs",
     "Date on which the swap begins accruing interest"),
    (None,"maturity_date","Maturity / Termination Date","YYYY-MM-DD","—","Both legs",
     "Date on which the final coupon and any principal exchange are paid;\nmust be strictly after start_date"),
    (None,"deal_date","Trade Date","YYYY-MM-DD","—","Both legs",
     "Date the swap was agreed; typically ~2 business days before start_date"),
    (None,"netting_id","Netting Agreement ID","String matching Netting_Database.csv","—  (required)","Both legs",
     "Groups trades under the same ISDA netting agreement for the Netting output;\nrequired — blank rows are excluded from the Netting CSV"),

    ("SCHEDULE  —  ROLL & GENERATION CONVENTIONS  (per leg: fixed_ / floating_ prefix)",None,None,None,None,None),
    (None,"*_roll_convention","Schedule Generation Direction",
     "forward\nforward_eom\nbackward\nbackward_eom","forward_eom","Both legs",
     "Controls how period boundaries are generated:\n"
     "forward: count from effective date; short stub at the end\n"
     "backward: count back from maturity; short stub at the front\n"
     "*_eom: if the anchor date is the last calendar day of its month, every subsequent boundary also snaps to month-end\n"
     "  (e.g. 31 Jan anchor -> all boundaries land on the last day of their month)"),
    (None,"*_bus_day_adj","Business Day Adjustment",
     "Following\nModifiedFollowing\nPreceding\nModifiedPreceding\nNearest\nNoAdjust",
     "required","Both legs",
     "When a period boundary or payment date falls on a weekend or holiday, this rule picks which business day to use:\n"
     "Following: next business day\n"
     "ModifiedFollowing: next business day, unless it crosses into the next calendar month — then go back  (most common for swaps)\n"
     "Preceding: prior business day\n"
     "NoAdjust: leave the date unchanged even if it is a holiday"),
    (None,"*_eff_date_adj","Effective Date Adjustment",
     "Same roll values or blank","Inherits *_bus_day_adj","Both legs",
     "Roll rule applied specifically to the effective (start) date.\nLeave blank to use the leg's *_bus_day_adj"),
    (None,"*_pay_date_adj","Payment Date Adjustment",
     "Same roll values or blank","Inherits *_bus_day_adj","Both legs",
     "Roll rule applied to payment dates.\nLeave blank to inherit the leg's *_bus_day_adj"),
    (None,"*_adjust","Accrual Basis",
     "acc_and_pay\npay\nnone","acc_and_pay","Both legs",
     "Determines which date boundaries the day-count fraction uses:\n"
     "acc_and_pay: day-count runs between ADJUSTED period boundaries  (Bloomberg default for most swap types)\n"
     "pay: day-count runs between UNADJUSTED boundaries; only the payment date is adjusted  (correct for 30/360)\n"
     "none: no adjustment applied  — test use only"),

    ("FIXED LEG  —  Conventions specific to the fixed coupon side",None,None,None,None,None),
    (None,"fixed_frequency","Fixed Coupon Frequency",
     "1Y / 6M / 3M / 1M / 1W / 1D","—","Fixed leg",
     "How often fixed coupons are paid.\nAnnual (1Y) and semi-annual (6M) are most common for USD IRS"),
    (None,"fixed_daycount","Fixed Day-Count Convention",
     "ACT/360\nACT/365F\n30/360\n30E/360\nACT/ACT-ISDA","ACT/360","Fixed leg",
     "Rule for converting a period from calendar days into a year fraction:\n"
     "ACT/360: actual days / 360  (money-market standard)\n"
     "ACT/365F: actual days / 365  (UK/AUS)\n"
     "30/360: assumes every month has 30 days and every year 360  (bond market)\n"
     "ACT/ACT-ISDA: actual days / actual days in the year  (government bonds)"),
    (None,"fixed_payment_delay_bdays","Fixed Payment Delay",
     "Integer >= 0","0","Fixed leg",
     "Business days after the accrual period ends before the coupon is paid.\n"
     "T+N is counted from the ADJUSTED period end (Bloomberg/ISDA standard).\n"
     "0 = pay on the last day of the adjusted period"),
    (None,"fixed_calculation_calendar","Fixed Accrual Calendar",
     "NY_FED  (currently only supported)","NY_FED","Fixed leg",
     "Holiday calendar used to determine business days for rolling period boundaries.\nCustom holidays can be added via fixed_calculation_calendar_extras or _extras_file"),
    (None,"fixed_payment_calendar","Fixed Payment Calendar",
     "NY_FED or blank","Inherits fixed_calculation_calendar","Fixed leg",
     "Calendar used when rolling payment dates.\nLeave blank to use the same calendar as accrual"),
    (None,"fixed_principal_exchange","Fixed Principal Exchange",
     "none / start / end / both","none","Fixed leg",
     "Whether to include notional flows:\n"
     "none: interest-only  (standard for vanilla swaps)\n"
     "start: pay out notional on the effective date\n"
     "end: receive notional back at maturity\n"
     "both: notional flows at both ends  (cross-currency style)"),
    (None,"fixed_first_period_accrual_end_date","First Payment Date Override",
     "YYYY-MM-DD or blank","blank  (auto-generated)","Fixed leg",
     "Forces the first accrual period to end on this date, creating a custom front stub.\n"
     "All subsequent periods then roll regularly from this anchor.\n"
     "Equivalent to Bloomberg SWPM 'First Payment Date'"),

    ("FLOATING LEG  —  Conventions specific to the Fed Funds floating side",None,None,None,None,None),
    (None,"floating_frequency","Floating Reset Frequency",
     "1Y / 6M / 3M / 1M / 1W / 1D  or blank","Inherits fixed_frequency","Floating leg",
     "How often the compounding period resets.\nFor Fed Funds OIS, typically matches the fixed leg.\nLeave blank to copy fixed_frequency"),
    (None,"floating_daycount","Floating Day-Count Convention",
     "ACT/360\nACT/365F\n30/360\n30E/360\nACT/ACT-ISDA","ACT/360","Floating leg",
     "Day-count for annualising the compounded rate.\nFor EFFR OIS always ACT/360  (matches the overnight compounding formula)"),
    (None,"floating_spread","Floating Spread",
     "Decimal  e.g. 0.0010 = 10bps\n0 for most vanilla swaps","0.0","Floating leg",
     "Basis-point spread added linearly to the compounded overnight rate each period.\nDoes not itself compound"),
    (None,"floating_lockout_bdays","Lockout Period",
     "Integer >= 0","0","Floating leg",
     "Freezes the last N overnight fixings at the rate observed N+1 business days before period end.\n"
     "Used when the settlement amount must be known before the last few fixings are published.\n"
     "Example: lockout=2 means the rate on the last 2 business days of the period is set to the rate from 3 days before period end"),
    (None,"floating_reset_lag_bdays","Fixing Lookback Lag",
     "Integer >= 0","0","Floating leg",
     "Shifts each fixing observation date backward by N business days.\n"
     "0 = in-arrears  (standard for EFFR OIS: rate for day T is the overnight rate published on T).\n"
     "Non-zero implements a lookback/observation-shift convention"),
    (None,"floating_rst_bus_day_adj","Fixing Roll Convention",
     "Same roll values or blank","Inherits floating_bus_day_adj","Floating leg",
     "Business-day roll applied to each fixing observation date after any lookback shift.\n"
     "With lag=0 this is a no-op because observation dates are already business days"),
    (None,"floating_fixing_calendar","Fixing Calendar",
     "NY_FED or blank","NY_FED","Floating leg",
     "Holiday calendar determining valid fixing publication dates.\n"
     "For EFFR always NY_FED  (NY Fed publishes the rate every NY business day)"),
    (None,"floating_payment_delay_bdays","Floating Payment Delay",
     "Integer >= 0","0","Floating leg",
     "Business days after the compounding period ends before the floating coupon is settled.\n"
     "T+N is counted from the ADJUSTED period end (Bloomberg/ISDA standard), then rolled by floating_pay_date_adj.\n"
     "Example: period ends Sunday May 10 -> adjusts to Monday May 11 -> T+2 = Wednesday May 13"),
    (None,"floating_principal_exchange","Floating Principal Exchange",
     "none / start / end / both","none","Floating leg",
     "Same as the fixed-leg equivalent.\nFor standard interest-only swaps, leave as none on both legs"),
    (None,"floating_first_period_accrual_end_date","Floating First Payment Date Override",
     "YYYY-MM-DD or blank","blank","Floating leg",
     "Same as fixed_first_period_accrual_end_date but for the floating schedule.\n"
     "Usually left blank so both legs share the same period structure"),

    ("PRODUCTION OUTPUT  —  Fields required for the KPMG regulatory CSV feed",None,None,None,None,None),
    (None,"quantum_deal_number","Quantum System Deal ID","String","—","Prod CSV col C",
     "Internal deal reference from the Quantum trade capture system"),
    (None,"oracle_entity_code","Oracle Entity Code","String  e.g. '1000', '1021'","—","Prod CSV cols D, AU, AV",
     "Legal entity identifier; used to look up the RC segment for CCID construction"),
    (None,"notional_currency","Currency","String  e.g. 'USD'","—","Prod CSV col E",
     "Notional currency; all trades in this book are USD"),
    (None,"intercompany","Intercompany Flag","true / false / yes / no / 1 / 0","—","Prod CSV col X",
     "Whether the trade is between two legal entities within the same group"),
    (None,"current_counterparty","Counterparty Name",
     "String\n'CME Clearing House' triggers CCP branch","—","Prod CSV cols Z, AE, AH, AN-AP",
     "Exact legal name of the counterparty.\n"
     "The exact string 'CME Clearing House' (case-sensitive, no extra spaces) routes the trade into the centrally-cleared branch, "
     "changing Sub-Product2, Counterparty Type, and the three CCP flags"),
    (None,"hedge","Hedge Direction","Long / Short","—  (required)","Prod CSV col AW",
     "Drives the Hedged Debt MTM column:\n"
     "Short: AW = negative of the swap's own clean value\n"
     "Long: AW = the hedged debt's Clean + USD Outstanding, looked up via Deal_Numbers.csv and Deal_Summary"),
    (None,"netting_id","Netting Agreement ID","String matching Netting_Database.csv","—  (required)","Prod CSV col AR; Netting CSV",
     "Key into the Netting Database; determines netting entity, position netting eligibility, and CCID segments for the Netting output"),
]

r = 3
for row in ROWS:
    is_sec = row[1] is None
    if is_sec:
        ws1.merge_cells(f"A{r}:F{r}")
        sec(ws1.cell(r, 1), row[0])
        ws1.row_dimensions[r].height = 18
    else:
        _, field, name, values, default, applies, impact = row
        alt = (r % 2 == 0)
        body(ws1.cell(r, 1), field,   alt=alt, bold=True)
        body(ws1.cell(r, 2), name,    alt=alt)
        body(ws1.cell(r, 3), values,  alt=alt)
        body(ws1.cell(r, 4), default, alt=alt)
        body(ws1.cell(r, 5), applies, alt=alt, align="center")
        body(ws1.cell(r, 6), impact,  alt=alt)
        ws1.row_dimensions[r].height = 72
    r += 1

# ════════════════════════════════════════════════════════════════════════
# TAB 2  Convention Examples
# ════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Convention Examples")
ws2.sheet_view.showGridLines = False
for i, w in enumerate([24, 22, 18, 18, 18, 18, 18, 44], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

banner(ws2, 1, 8, "Swap Pricer — Convention Examples  (Concrete Date Illustrations)")
ws2.row_dimensions[1].height = 28

# Example 1 — roll_convention
ex_title(ws2, 3, 8, "Example 1 — *_roll_convention:  forward  vs  backward  vs  forward_eom")
note(ws2, 4, 8,
     "Trade: Quarterly (3M), effective 31 Jan 2026, maturity 31 Oct 2026.  "
     "forward_eom snaps to month-end because 31 Jan is the last day of January.")
for c, h in enumerate(["roll_convention","Period 1  Start -> End","Period 2  Start -> End",
                        "Period 3  Start -> End","Stub?","","","Key Takeaway"], 1):
    hdr(ws2.cell(5, c), h)
ws2.row_dimensions[5].height = 32

ex1 = [
    ("forward",
     "31 Jan -> 30 Apr","30 Apr -> 31 Jul","31 Jul -> 30 Oct","None (full periods)","","",
     "Counts forward. Period ends stay at the calendar date without month-end snapping."),
    ("forward_eom",
     "31 Jan -> 30 Apr","30 Apr -> 31 Jul","31 Jul -> 31 Oct","None (full periods)","","",
     "Same as forward BUT every boundary snaps to the last day of its month because 31 Jan is month-end.\nNotice Oct ends on 31st, not 30th."),
    ("backward",
     "31 Jan -> 30 Apr","30 Apr -> 31 Jul","31 Jul -> 31 Oct","None (full periods)","","",
     "Counts back from 31 Oct. For this trade the dates coincide with forward,\nbut for irregular maturities backward will produce different period starts."),
    ("backward_eom",
     "31 Jan -> 28 Feb","28 Feb -> 31 May","31 May -> 31 Aug","Short front stub:\n31 Jan -> 30 Apr splits","","",
     "Backward + month-end: the stub (leftover days) absorbs at the FRONT of the schedule."),
]
fills_1 = [WHT_FILL, YLW_FILL, ALT_FILL, WHT_FILL]
for i, (row, fill) in enumerate(zip(ex1, fills_1)):
    for c, val in enumerate(row, 1):
        body(ws2.cell(6+i, c), val, fill=fill, bold=(c == 1))
    ws2.row_dimensions[6+i].height = 52

# Example 2 — *_adjust
ex_title(ws2, 11, 8, "Example 2 — *_adjust:  acc_and_pay  vs  pay  (effect on day-count fraction)")
note(ws2, 12, 8,
     "Period: unadjusted 28 Feb -> 31 May.  "
     "28 Feb (Saturday) adjusts to 2 Mar under ModifiedFollowing.  "
     "31 May (Sunday) adjusts to 1 Jun.  Day-count convention: 30/360.")
for c, h in enumerate(["adjust","Accrual Start\n(for DCF)","Accrual End\n(for DCF)",
                        "Day-Count Fraction\n(30/360)","Payment Date\n(always adjusted)","","","Key Takeaway"], 1):
    hdr(ws2.cell(13, c), h)
ws2.row_dimensions[13].height = 36

ex2 = [
    ("acc_and_pay","2 Mar  (adjusted)","1 Jun  (adjusted)","90/360 = 0.25000","1 Jun","","",
     "Day-count uses adjusted dates. Standard for ACT/360 swaps — the accrual period matches what was actually observed."),
    ("pay","28 Feb  (unadjusted)","31 May  (unadjusted)","93/360 = 0.25833","1 Jun","","",
     "Day-count uses the theoretical dates. Payment is still adjusted.\nCorrect for 30/360 swaps where the convention assumes clean calendar boundaries."),
]
fills_2 = [GRN_FILL, YLW_FILL]
for i, (row, fill) in enumerate(zip(ex2, fills_2)):
    for c, val in enumerate(row, 1):
        body(ws2.cell(14+i, c), val, fill=fill, bold=(c == 1))
    ws2.row_dimensions[14+i].height = 52

# Example 3 — lockout
ex_title(ws2, 17, 8, "Example 3 — floating_lockout_bdays:  freezing the last N fixing rates")
note(ws2, 18, 8,
     "Period ends Friday 28 Mar.  lockout_bdays = 2.  "
     "The last 2 business days of the period (Thu 27 Mar, Wed 26 Mar) reuse the rate from Tue 25 Mar.")
for c, h in enumerate(["Fixing Date","Normal Rate\n(curve / history)","Applied Rate\n(lockout = 2)",
                        "Rate Source Tag","","","","Notes"], 1):
    hdr(ws2.cell(19, c), h)
ws2.row_dimensions[19].height = 36

ex3 = [
    ("... earlier days ...","varies","varies","history / curve","","","",""),
    ("Tue 25 Mar","5.3300%","5.3300%","curve","","","","Last normal fixing"),
    ("Wed 26 Mar","5.3350%","5.3300%","lockout","","","","Frozen at Tue's rate"),
    ("Thu 27 Mar","5.3400%","5.3300%","lockout","","","","Frozen at Tue's rate"),
]
fills_3 = [WHT_FILL, GRN_FILL, YLW_FILL, YLW_FILL]
for i, (row, fill) in enumerate(zip(ex3, fills_3)):
    for c, val in enumerate(row, 1):
        body(ws2.cell(20+i, c), val, fill=fill, bold=(c in (1, 4)))
    ws2.row_dimensions[20+i].height = 30

# Example 4 — payment delay
ex_title(ws2, 25, 8, "Example 4 — *_payment_delay_bdays:  how the settlement date shifts")
note(ws2, 26, 8,
     "Accrual period ends Monday 30 Jun 2026.  "
     "Delay measured in business days from the unadjusted period end, then rolled by Pay Date Adj (ModifiedFollowing).")
for c, h in enumerate(["payment_delay_bdays","Unadjusted Period End","Settlement Date",
                        "Notes","","","",""], 1):
    hdr(ws2.cell(27, c), h)
ws2.row_dimensions[27].height = 30

ex4 = [
    ("0  (default)","30 Jun 2026","30 Jun 2026","Settle on the last day of the period — no lag"),
    ("1","30 Jun 2026","1 Jul 2026","T+1: one business day after period end"),
    ("2","30 Jun 2026","2 Jul 2026","T+2: standard for many USD OIS structures"),
    ("5","30 Jun 2026","7 Jul 2026","T+5: settle one week later (skips 4 Jul holiday if applicable)"),
]
for i, row in enumerate(ex4):
    alt = (i % 2 == 1)
    vals = list(row) + ["","","",""]
    for c, val in enumerate(vals, 1):
        body(ws2.cell(28+i, c), val, alt=alt, bold=(c == 1))
    ws2.row_dimensions[28+i].height = 28

# ════════════════════════════════════════════════════════════════════════
# TAB 3  Inheritance Rules
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Inheritance Rules")
ws3.sheet_view.showGridLines = False
for i, w in enumerate([30, 30, 30, 44, 26], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

banner(ws3, 1, 5, "Swap Pricer — Field Inheritance & Auto-Sync Rules")
note(ws3, 2, 5,
     "When a field is left BLANK in the trade CSV/YAML, the pricer automatically fills it from another field (shown below).  "
     "You only need to specify a value when you want behaviour that differs from the default.")
ws3.row_dimensions[2].height = 44
ws3.merge_cells("A2:E2")
ws3.freeze_panes = "A3"

for c, h in enumerate(["Field Left Blank","Inherits From","Inherited Value (example)",
                        "When to Override","Affected Output"], 1):
    hdr(ws3.cell(3, c), h)
ws3.row_dimensions[3].height = 30

inh = [
    ("CALENDAR INHERITANCE",None,None,None,None),
    ("fixed_payment_calendar","fixed_calculation_calendar","NY_FED",
     "If fixed coupons settle on a different holiday calendar from the accrual calendar (rare for USD swaps)",
     "Fixed leg payment dates"),
    ("floating_payment_calendar","floating_calculation_calendar","NY_FED",
     "If floating coupons settle on a different calendar","Floating leg payment dates"),
    ("floating_fixing_calendar","floating_calculation_calendar","NY_FED",
     "If the index publication calendar differs from the period-boundary calendar (not the case for EFFR)",
     "Floating fixing observation dates"),

    ("ROLL CONVENTION INHERITANCE",None,None,None,None),
    ("fixed_eff_date_adj","fixed_bus_day_adj","ModifiedFollowing",
     "When the effective date must roll under a different rule  (e.g. NoAdjust if locked to a specific date)",
     "Effective date on fixed leg"),
    ("fixed_pay_date_adj","fixed_bus_day_adj","ModifiedFollowing",
     "When payment dates must use a different roll from accrual boundaries  (uncommon)",
     "Fixed payment dates"),
    ("floating_eff_date_adj","floating_bus_day_adj","ModifiedFollowing",
     "Same as fixed — only set if the floating effective date needs its own rule",
     "Effective date on floating leg"),
    ("floating_pay_date_adj","floating_bus_day_adj","ModifiedFollowing",
     "When floating payment dates use a different roll from accrual boundaries",
     "Floating payment dates"),
    ("floating_rst_bus_day_adj","floating_bus_day_adj","ModifiedFollowing",
     "When fixing observation dates (after any lookback shift) must roll differently.\nWith lag=0 this is a no-op because dates are already business days",
     "Daily fixing observation dates"),

    ("FREQUENCY INHERITANCE",None,None,None,None),
    ("floating_frequency","fixed_frequency","Matches fixed leg  e.g. 3M",
     "When the floating leg resets on a different schedule from the fixed coupon payments  (unusual for vanilla OIS)",
     "Floating accrual period boundaries"),

    ("PRODUCTION OUTPUT DEFAULTS — sourced from external files, never the trade CSV",None,None,None,None),
    ("cash_flow_netting_allowed","Netting_Database.csv  keyed by netting_id","Lookup value",
     "Never override in the trade file — always sourced from the centralised netting database",
     "Prod CSV col AS;  Netting CSV col Q"),
    ("position_netting_allowed","Netting_Database.csv  keyed by netting_id","Lookup value",
     "Never override in the trade file — always sourced from the centralised netting database",
     "Prod CSV col AT;  Netting CSV col R"),
    ("oracle_entity_code (RC segment)","Entity_Reference_Report.csv  keyed by oracle_entity_code","Lookup value",
     "The RC is always looked up from the entity report — it is not set on the trade directly",
     "Prod CSV cols AU, AV (CCID);  Netting CSV cols T, U"),
]

r3 = 4
for row in inh:
    is_sec = row[1] is None
    if is_sec:
        ws3.merge_cells(f"A{r3}:E{r3}")
        sec(ws3.cell(r3, 1), row[0])
        ws3.row_dimensions[r3].height = 18
    else:
        alt = (r3 % 2 == 0)
        for c, val in enumerate(row, 1):
            body(ws3.cell(r3, c), val, alt=alt, bold=(c == 1))
        ws3.row_dimensions[r3].height = 60
    r3 += 1

wb.save("Convention_Reference.xlsx")
print("Saved: Convention_Reference.xlsx")
